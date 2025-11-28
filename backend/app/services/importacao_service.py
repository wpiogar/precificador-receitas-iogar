# ============================================================================
# SERVICE - PROCESSAMENTO DE IMPORTAÇÃO DE INSUMOS
# ============================================================================
# Descrição: Serviço para processar arquivos Excel e importar insumos
# Data: 30/10/2025
# Autor: Will - Empresa: IOGAR
# ============================================================================

import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import func
import json
import logging

from app.models.importacao_insumo import ImportacaoInsumo, StatusImportacao
from app.models.insumo import Insumo
from app.schemas.importacao_insumo import (
    PreviewImportacao,
    LogProcessamento,
    ItemLog
)

# Configurar logging
logger = logging.getLogger(__name__)


# ============================================================================
# CONSTANTES DE MAPEAMENTO
# ============================================================================

MAPEAMENTO_COLUNAS = {
    'CodigoProduto': 'codigo',
    'Código': 'codigo',  
    'NomeProduto': 'nome',
    'Produto': 'nome',  
    'PrecoCompra': 'preco_compra_real',
    'Preço Compra': 'preco_compra_real',  
    'Unidade': 'unidade',
    'Un': 'unidade',  
    'Fator': 'fator',
    'Quantidade': 'quantidade'
}

CONVERSAO_UNIDADES = {
    'LT': 'L',
    'UN': 'unidade',
    'KG': 'kg',
    'G': 'g',
    'ML': 'ml',
    'L': 'L'
}


# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

def converter_unidade(unidade_excel: str) -> str:
    """
    Converte unidade do Excel para o padrão do sistema.
    
    Args:
        unidade_excel: Unidade vinda do Excel (LT, UN, KG, etc.)
        
    Returns:
        str: Unidade no padrão do sistema
        
    Exemplos:
        LT -> L
        UN -> unidade
        KG -> kg
    """
    if not unidade_excel:
        return 'unidade'
    
    unidade_upper = str(unidade_excel).strip().upper()
    return CONVERSAO_UNIDADES.get(unidade_upper, unidade_excel.lower())


def encontrar_linha_cabecalho(worksheet) -> Optional[int]:
    """
    Encontra a linha do cabeçalho no Excel.
    Identifica o cabeçalho pela primeira linha que contém múltiplos valores não vazios
    e não é apenas numérica (indicando dados).
    
    Args:
        worksheet: Planilha do openpyxl
        
    Returns:
        int: Número da linha do cabeçalho (1-indexed) ou None se não encontrar
    """
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20), start=1):
        valores = [str(cell.value).strip() if cell.value else '' for cell in row]
        
        # Contar quantos valores não vazios existem
        valores_preenchidos = [v for v in valores if v and v.lower() not in ['none', 'nan', '']]
        
        # Se tiver pelo menos 3 valores preenchidos, pode ser o cabeçalho
        if len(valores_preenchidos) >= 3:
            # Verificar se não é uma linha de dados (valores numéricos)
            # Cabeçalho geralmente tem pelo menos um texto
            tem_texto = any(
                not v.replace('.', '').replace(',', '').replace('-', '').isdigit() 
                for v in valores_preenchidos
            )
            
            if tem_texto:
                logger.info(f"Cabeçalho encontrado na linha {row_num}: {valores_preenchidos[:5]}")
                return row_num
    
    # Se não encontrou cabeçalho nas primeiras 20 linhas, assume linha 1
    logger.warning("Cabeçalho não identificado automaticamente, usando linha 1")
    return 1

def encontrar_coluna_real_com_dados(row, idx_inicial: int, validador=None) -> int:
    """
    Quando uma coluna está mesclada, encontra qual coluna tem os dados reais.
    
    Args:
        row: Linha do Excel
        idx_inicial: Índice inicial da coluna mapeada
        validador: Função opcional para validar se o valor é válido
        
    Returns:
        int: Índice da coluna que contém os dados reais
        
    Exemplos:
        - Cabeçalho "Código" mesclado em A-B-C, mas dados em C
        - Retorna índice 2 (coluna C)
    """
    # Converter row em lista
    if not isinstance(row, (list, tuple)):
        row = list(row)
    
    # Tentar até 3 colunas à frente do índice inicial
    for offset in range(4):  # Tenta idx, idx+1, idx+2, idx+3
        idx_teste = idx_inicial + offset
        
        if idx_teste >= len(row):
            break
        
        cell = row[idx_teste]
        valor = cell.value if hasattr(cell, 'value') else cell
        
        # Ignorar valores vazios
        if not valor or str(valor).strip() in ['', 'None', 'none']:
            continue
        
        # Se tem validador, usar ele
        if validador:
            if validador(valor):
                logger.info(f"Coluna real encontrada: índice {idx_teste} (offset +{offset})")
                return idx_teste
        else:
            # Sem validador, retornar primeiro valor não vazio
            logger.info(f"Coluna real encontrada: índice {idx_teste} (offset +{offset})")
            return idx_teste
    
    # Se não encontrou, retornar índice original
    return idx_inicial

def validar_codigo_insumo(valor) -> bool:
    """
    Valida se um valor é um código de insumo válido (numérico entre 5000-7999).
    
    Args:
        valor: Valor a validar
        
    Returns:
        bool: True se for código válido, False caso contrário
    """
    try:
        # Converter para string e limpar
        codigo_str = str(valor).strip()
        
        # Remover .0 se houver (ex: 5228.0 -> 5228)
        if '.' in codigo_str:
            codigo_float = float(codigo_str)
            if codigo_float == int(codigo_float):
                codigo_str = str(int(codigo_float))
        
        # Converter para inteiro
        codigo_num = int(codigo_str)
        
        # Validar faixa
        return 5000 <= codigo_num <= 7999
        
    except (ValueError, TypeError):
        return False


def extrair_dados_linha(row, colunas_mapeadas: Dict[str, int]) -> Dict[str, Any]:
    """
    Extrai dados de uma linha do Excel usando o mapeamento de colunas.
    
    Args:
        row: Linha do Excel (tupla de células)
        colunas_mapeadas: Dicionário {nome_campo: índice_coluna}
        
    Returns:
        Dict com os dados extraídos e convertidos
    """
    dados = {}
    
    # Converter row em lista se for iterator
    if not isinstance(row, (list, tuple)):
        row = list(row)
    
    # ====================================================================
    # Extrair código - LIMPAR formato decimal do Excel (5228.0 -> 5228)
    # Detecta automaticamente a coluna real em caso de células mescladas
    # ====================================================================
    if 'codigo' in colunas_mapeadas:
        idx_original = colunas_mapeadas['codigo']
        
        # Encontrar coluna real com dados válidos (células mescladas)
        idx = encontrar_coluna_real_com_dados(row, idx_original, validador=validar_codigo_insumo)
        
        if idx < len(row):
            codigo_cell = row[idx]
            valor = codigo_cell.value if hasattr(codigo_cell, 'value') else codigo_cell
            if valor:
                codigo_str = str(valor).strip()
                # Se o código é numérico com .0, remover o decimal
                if '.' in codigo_str:
                    try:
                        # Tentar converter para float e depois para int
                        codigo_float = float(codigo_str)
                        # Se é um número inteiro (ex: 5228.0), converter para int
                        if codigo_float == int(codigo_float):
                            dados['codigo'] = str(int(codigo_float))
                        else:
                            dados['codigo'] = codigo_str
                    except (ValueError, TypeError):
                        dados['codigo'] = codigo_str
                else:
                    dados['codigo'] = codigo_str
            else:
                dados['codigo'] = None
    
    # Extrair nome
    if 'nome' in colunas_mapeadas:
        idx = colunas_mapeadas['nome']
        if idx < len(row):
            nome_cell = row[idx]
            valor = nome_cell.value if hasattr(nome_cell, 'value') else nome_cell
            dados['nome'] = str(valor).strip() if valor else None
    
    # Extrair preço - verificar preco_unitario E preco_compra_real
    preco_key = None
    if 'preco_unitario' in colunas_mapeadas:
        preco_key = 'preco_unitario'
    elif 'preco_compra_real' in colunas_mapeadas:
        preco_key = 'preco_compra_real'
    
    if preco_key:
        idx = colunas_mapeadas[preco_key]
        if idx < len(row):
            preco_cell = row[idx]
            valor = preco_cell.value if hasattr(preco_cell, 'value') else preco_cell
            try:
                preco = float(valor) if valor else 0.0
                dados['preco_compra_real'] = preco
            except (ValueError, TypeError):
                dados['preco_compra_real'] = 0.0
    
    # Extrair unidade
    if 'unidade' in colunas_mapeadas:
        idx = colunas_mapeadas['unidade']
        if idx < len(row):
            unidade_cell = row[idx]
            valor = unidade_cell.value if hasattr(unidade_cell, 'value') else unidade_cell
            unidade_raw = str(valor).strip() if valor else 'unidade'
            dados['unidade'] = converter_unidade(unidade_raw)
    
    # Extrair quantidade (se mapeado)
    if 'fator' in colunas_mapeadas:
        idx = colunas_mapeadas['fator']
        if idx < len(row):
            fator_cell = row[idx]
            valor = fator_cell.value if hasattr(fator_cell, 'value') else fator_cell
            try:
                dados['fator'] = float(valor) if valor else 1.0
            except (ValueError, TypeError):
                dados['fator'] = 1.0
    
    # Extrair grupo/categoria (se mapeado)
    if 'grupo' in colunas_mapeadas:
        idx = colunas_mapeadas['grupo']
        if idx < len(row):
            grupo_cell = row[idx]
            valor = grupo_cell.value if hasattr(grupo_cell, 'value') else grupo_cell
            dados['grupo'] = str(valor).strip() if valor else None
    
    return dados


def validar_dados_linha(dados: Dict[str, Any], linha_num: int) -> Tuple[bool, Optional[str]]:
    """
    Valida se os dados extraídos estão corretos.
    
    Args:
        dados: Dicionário com dados extraídos
        linha_num: Número da linha (para mensagem de erro)
        
    Returns:
        Tuple[bool, Optional[str]]: (válido, mensagem_erro)
    """
    # Validar código (obrigatório na importação)
    codigo = dados.get('codigo')
    if not codigo or str(codigo).strip() == '' or str(codigo).lower() in ['none', 'nan']:
        return False, f"Linha {linha_num}: Código do produto está vazio ou inválido"
    
    # Validar nome (obrigatório)
    nome = dados.get('nome')
    if not nome or str(nome).strip() == '' or str(nome).lower() in ['none', 'nan']:
        return False, f"Linha {linha_num}: Nome do produto está vazio"
    
    # Unidade - se não tiver, usar padrão
    if not dados.get('unidade'):
        dados['unidade'] = 'unidade'
    
    # Preço - se não tiver, usar 0
    if 'preco_compra_real' not in dados or dados['preco_compra_real'] is None:
        dados['preco_compra_real'] = 0.0
    
    return True, None


# ============================================================================
# CLASSE PRINCIPAL DO SERVICE
# ============================================================================

class ImportacaoService:
    """
    Service para processamento de importação de insumos via Excel.
    """
    
    def __init__(self, db: Session):
        """
        Inicializa o service com a sessão do banco de dados.
        
        Args:
            db: Sessão do SQLAlchemy
        """
        self.db = db
    
    # ========================================================================
    # MÉTODO: GERAR PREVIEW
    # ========================================================================
    
    def gerar_preview(
        self,
        caminho_arquivo: str,
        nome_arquivo: str
    ) -> PreviewImportacao:
        """
        Gera preview dos dados que serão importados.
        
        Args:
            caminho_arquivo: Caminho do arquivo Excel no servidor
            nome_arquivo: Nome original do arquivo
            
        Returns:
            PreviewImportacao: Schema com preview dos dados
            
        Raises:
            ValueError: Se arquivo inválido ou não puder ser lido
        """
        try:
            # Abrir arquivo Excel
            wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
            ws = wb.active
            
            # Encontrar linha do cabeçalho
            linha_cabecalho = encontrar_linha_cabecalho(ws)
            if not linha_cabecalho:
                raise ValueError("Não foi possível encontrar o cabeçalho no arquivo Excel")
            
            # Extrair colunas do cabeçalho - TODAS AS COLUNAS (mesmo com vazias no meio)
            header_row = list(ws.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho))[0]
            colunas_detectadas = []

            # Detectar até onde tem dados (última coluna com conteúdo)
            ultima_coluna_com_dado = 0
            for i, cell in enumerate(header_row):
                if cell.value is not None and str(cell.value).strip() != '':
                    ultima_coluna_com_dado = i

            # Extrair até a última coluna que tem dado
            for i in range(ultima_coluna_com_dado + 1):
                cell = header_row[i]
                valor = cell.value
                if valor is None or str(valor).strip() == '':
                    # Coluna vazia - dar nome genérico
                    colunas_detectadas.append(f"Coluna_{i+1}")
                else:
                    colunas_detectadas.append(str(valor).strip())

            logger.info(f"📋 Colunas detectadas no preview ({len(colunas_detectadas)}): {colunas_detectadas}")
            
            # Mapear colunas
            colunas_mapeadas = {}
            for i, col in enumerate(colunas_detectadas):
                campo_sistema = MAPEAMENTO_COLUNAS.get(col)
                if campo_sistema:
                    colunas_mapeadas[campo_sistema] = i
            
            # Extrair primeiras 5 linhas de dados
            primeiras_linhas = []
            linha_dados_inicial = linha_cabecalho + 1
            
            for row in ws.iter_rows(min_row=linha_dados_inicial, max_row=linha_dados_inicial + 4):
                dados = extrair_dados_linha(row, colunas_mapeadas)
                if dados.get('codigo') and dados.get('nome'):
                    primeiras_linhas.append(dados)
            
            # Contar total de linhas
            total_linhas = ws.max_row - linha_cabecalho
            
            # Gerar avisos
            avisos = []
            if 'codigo' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'CodigoProduto' não encontrada")
            if 'nome' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'NomeProduto' não encontrada")
            if 'preco_compra_real' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'PrecoCompra' não encontrada")
            if 'unidade' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'Unidade' não encontrada")
            
            # ================================================================
            # Informar sobre coluna Fator (opcional)
            # ================================================================
            # HISTÓRICO: Aviso adicionado em 19/11/2025
            if 'quantidade' in colunas_mapeadas:
                avisos.append("ℹ️ Coluna 'Fator' detectada - valores serão importados como Quantidade")
            
            wb.close()
            
            return PreviewImportacao(
                nome_arquivo=nome_arquivo,
                total_linhas=total_linhas,
                colunas_detectadas=colunas_detectadas,
                primeiras_linhas=primeiras_linhas,
                mapeamento_colunas={
                    k: v for k, v in MAPEAMENTO_COLUNAS.items()
                },
                avisos=avisos
            )
            
        except Exception as e:
            logger.error(f"Erro ao gerar preview: {e}")
            raise ValueError(f"Erro ao processar arquivo: {str(e)}")
    
    # ========================================================================
    # MÉTODO: PROCESSAR IMPORTAÇÃO
    # ========================================================================
    
    def processar_importacao(
        self,
        importacao_id: int,
        restaurante_id: int,
        mapeamento_customizado: Optional[Dict[str, str]] = None
    ) -> Tuple[bool, str]:
        """
        Processa a importação e cria os insumos no banco de dados.
        
        Args:
            importacao_id: ID da importação a processar
            restaurante_id: ID do restaurante
            
        Returns:
            Tuple[bool, str]: (sucesso, mensagem)
        """
        # Buscar importação
        importacao = self.db.query(ImportacaoInsumo).filter(
            ImportacaoInsumo.id == importacao_id
        ).first()
        
        if not importacao:
            return False, "Importação não encontrada"
        
        # Atualizar status para processando
        importacao.status = StatusImportacao.PROCESSANDO
        importacao.data_inicio_processamento = datetime.now()
        self.db.commit()
        
        # Inicializar log
        log = LogProcessamento()
        
        try:
            # Abrir arquivo Excel
            wb = openpyxl.load_workbook(importacao.caminho_arquivo, data_only=True)
            ws = wb.active
            
            # Encontrar cabeçalho
            linha_cabecalho = encontrar_linha_cabecalho(ws)
            if not linha_cabecalho:
                raise ValueError("Cabeçalho não encontrado")
            
            # Mapear colunas
            header_row = list(ws.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho))[0]
            colunas_mapeadas = {}

            # Se recebeu mapeamento customizado do frontend, usar ele
            if mapeamento_customizado:
                logger.info(f"Usando mapeamento customizado: {mapeamento_customizado}")
                for i, cell in enumerate(header_row):
                    col_name_excel = str(cell.value).strip() if cell.value else ''
                    # Verificar se esta coluna está no mapeamento customizado
                    if col_name_excel in mapeamento_customizado:
                        campo_sistema = mapeamento_customizado[col_name_excel]
                        colunas_mapeadas[campo_sistema] = i
                        logger.info(f"Mapeado: '{col_name_excel}' -> '{campo_sistema}' (coluna {i})")
            else:
                # Usar mapeamento padrão (TOTVS)
                logger.info("Usando mapeamento padrão TOTVS")
                for i, cell in enumerate(header_row):
                    col_name = str(cell.value).strip() if cell.value else ''
                    campo_sistema = MAPEAMENTO_COLUNAS.get(col_name)
                    if campo_sistema:
                        colunas_mapeadas[campo_sistema] = i

            # ADICIONAR ESTAS 2 LINHAS AQUI
            logger.info(f"📊 Colunas mapeadas final: {colunas_mapeadas}")
            logger.info(f"📋 Total de colunas no header: {len(header_row)}")            
            
            # Processar linhas
            linha_dados_inicial = linha_cabecalho + 1
            total_linhas = ws.max_row - linha_cabecalho

            logger.info(f"🚀 Iniciando processamento de {total_linhas} linhas a partir da linha {linha_dados_inicial}")

            for row_num, row in enumerate(
                ws.iter_rows(min_row=linha_dados_inicial),
                start=linha_dados_inicial
            ):
                try:
                    # Log início do processamento da linha
                    logger.info(f"🔄 Processando linha {row_num}")
                    
                    # Extrair dados
                    dados = extrair_dados_linha(row, colunas_mapeadas)
                    logger.info(f"📋 Dados extraídos linha {row_num}: {dados}")
                    
                    # VERIFICAÇÃO 1: Ignorar linhas de cabeçalho de grupo
                    codigo_valor = dados.get('codigo', '')
                    if codigo_valor and str(codigo_valor).strip().upper().startswith('GRUPO'):
                        logger.info(f"⏭️ Ignorando linha {row_num}: cabeçalho de grupo")
                        log.ignorados.append(ItemLog(
                            linha=row_num,
                            tipo="ignorado",
                            mensagem="Linha de cabeçalho de grupo",
                            dados={'codigo': codigo_valor}
                        ))
                        continue
                    
                    # VERIFICAÇÃO 2: Ignorar linhas completamente vazias
                    tem_algum_valor = any(
                        dados.get(campo) and str(dados.get(campo)).strip() not in ['', 'None', 'none']
                        for campo in ['codigo', 'nome']
                    )
                    if not tem_algum_valor:
                        logger.info(f"⏭️ Ignorando linha {row_num}: vazia")
                        log.ignorados.append(ItemLog(
                            linha=row_num,
                            tipo="ignorado",
                            mensagem="Linha vazia",
                            dados={}
                        ))
                        continue
                    
                    # VERIFICAÇÃO 3: Gerar código automático se necessário
                    if (not dados.get('codigo') or str(dados.get('codigo')).strip() in ['', 'None', 'none']) and dados.get('nome'):
                        # Importar serviço de geração de códigos
                        from app.services.codigo_service import gerar_proximo_codigo
                        from app.config.codigo_config import TipoCodigo
                        
                        # Gerar código usando sistema oficial (a partir de 5000, sem limite superior)
                        codigo_final = gerar_proximo_codigo(self.db, TipoCodigo.INSUMO, restaurante_id)
                        dados['codigo'] = codigo_final
                        logger.info(f"Linha {row_num}: Código gerado automaticamente: {codigo_final}")
                    
                    # FILTRO: Apenas códigos a partir de 5000 (sem limite superior)
                    try:
                        codigo_str = str(dados.get('codigo', '')).strip()
                        codigo_numero = int(codigo_str)
                        if codigo_numero < 5000:
                            log.ignorados.append(ItemLog(
                                linha=row_num,
                                tipo="ignorado",
                                mensagem=f"Código {dados['codigo']} abaixo do mínimo permitido (mínimo: 5000)",
                                dados=dados
                            ))
                            continue
                    except (ValueError, TypeError):
                        log.erros.append(ItemLog(
                            linha=row_num,
                            tipo="erro",
                            mensagem=f"Código inválido (não numérico): {dados.get('codigo')}",
                            dados=dados
                        ))
                        continue
                    
                    # Validar dados
                    valido, erro = validar_dados_linha(dados, row_num)
                    if not valido:
                        log.erros.append(ItemLog(
                            linha=row_num,
                            tipo="erro",
                            mensagem=erro,
                            dados=dados
                        ))
                        continue
                    
                    # ====================================================================
                    # IMPORTANTE: Limpar cache do SQLAlchemy antes de verificar duplicados
                    # Isso garante que após uma limpeza de insumos, a verificação
                    # consulta dados frescos do banco, não do cache da sessão
                    # ====================================================================
                    self.db.flush()
                    self.db.expire_all()

                    # Verificar se insumo já existe (consulta fresca no banco)
                    insumo_existente = self.db.query(Insumo).filter(
                        Insumo.restaurante_id == restaurante_id,
                        Insumo.codigo == dados['codigo']
                    ).first()

                    if insumo_existente:
                        log.ignorados.append(ItemLog(
                            linha=row_num,
                            tipo="ignorado",
                            mensagem=f"Código {dados['codigo']} já existe no sistema (insumo cadastrado: '{insumo_existente.nome}')",
                            dados=dados
                        ))
                        logger.info(f"⏭️ Pulando linha {row_num}: código {dados['codigo']} duplicado")
                        continue
                    
                    # ================================================================
                    # CRIAR INSUMO NO BANCO DE DADOS
                    # ================================================================
                    try:
                        logger.info(f"🆕 Tentando criar insumo linha {row_num}")
                        
                        # Preparar quantidade
                        quantidade = dados.get('quantidade', 1)
                        if quantidade is None or quantidade == 0:
                            quantidade = 1
                        
                        # Preparar preço
                        preco_real = dados.get('preco_unitario', dados.get('preco_compra_real', 0)) or 0
                        preco_centavos = int(float(preco_real) * 100) if preco_real else 0
                        
                        logger.info(f"📦 Valores preparados - codigo: {dados.get('codigo')}, nome: {dados.get('nome')}, qtd: {quantidade}, preco_centavos: {preco_centavos}")
                        
                        # Criar objeto Insumo
                        novo_insumo = Insumo(
                            restaurante_id=restaurante_id,
                            importacao_id=importacao_id,
                            codigo=dados['codigo'],
                            nome=dados['nome'],
                            quantidade=float(quantidade),
                            fator=float(dados.get('fator', 1.0)),  # Adicionar fator da importação
                            unidade=dados['unidade'],
                            preco_compra=preco_centavos,
                            grupo='',
                            subgrupo='',
                            eh_fornecedor_anonimo=True,
                            aguardando_classificacao=False
                        )
                        
                        logger.info(f"➕ Adicionando insumo ao banco: {dados['codigo']} - {dados['nome']}")
                        self.db.add(novo_insumo)
                        self.db.flush()  # Força commit imediato para detectar erros
                        logger.info(f"✅ Insumo linha {row_num} adicionado com sucesso ao banco")
                        
                        log.sucessos.append(ItemLog(
                            linha=row_num,
                            tipo="sucesso",
                            mensagem=f"Insumo '{dados['nome']}' importado",
                            dados=dados
                        ))
                        
                    except Exception as e_create:
                        # Log detalhado de erro na criação
                        import traceback
                        logger.error(f"❌ ERRO ao criar insumo linha {row_num}")
                        logger.error(f"   Código: {dados.get('codigo')}")
                        logger.error(f"   Nome: {dados.get('nome')}")
                        logger.error(f"   Erro: {str(e_create)}")
                        logger.error(f"   Tipo: {type(e_create).__name__}")
                        logger.error(f"   Traceback: {traceback.format_exc()}")
                        
                        log.erros.append(ItemLog(
                            linha=row_num,
                            tipo="erro",
                            mensagem=f"Erro ao criar insumo: {str(e_create)}",
                            dados=dados
                        ))
                        
                        # Rollback desta transação específica
                        self.db.rollback()
                        continue
                    
                except Exception as e_linha:
                    import traceback
                    logger.error(f"❌ ERRO linha {row_num}: {str(e_linha)}")
                    logger.error(f"   Traceback:\n{traceback.format_exc()}")
                    logger.error(f"   Dados: {dados if 'dados' in locals() else 'N/A'}")
                    
                    log.erros.append(ItemLog(
                        linha=row_num,
                        tipo="erro",
                        mensagem=f"Erro: {str(e_linha)}",
                        dados=dados if 'dados' in locals() else {}
                    ))

                # ============================================================
                # COMMIT EM LOTE - A cada 50 sucessos, salvar no banco
                # ============================================================
                if len(log.sucessos) > 0 and len(log.sucessos) % 50 == 0:
                    try:
                        self.db.commit()
                        logger.info(f"💾 Commit em lote: {len(log.sucessos)} insumos salvos no banco")
                    except Exception as e_commit:
                        logger.error(f"❌ Erro no commit em lote: {str(e_commit)}")
                        self.db.rollback()
                
                # Log de progresso a cada 50 linhas processadas
                if row_num % 50 == 0:
                    logger.info(f"📊 Progresso linha {row_num} - Sucessos: {len(log.sucessos)}, Erros: {len(log.erros)}, Ignorados: {len(log.ignorados)}")

            wb.close()

            # Atualizar estatísticas da importação
            importacao.total_linhas = total_linhas
            importacao.linhas_processadas = len(log.sucessos)
            importacao.linhas_com_erro = len(log.erros)
            importacao.linhas_ignoradas = len(log.ignorados)
            importacao.log_processamento = log.model_dump_json()
            importacao.data_fim_processamento = datetime.now()

            # Definir status final
            if len(log.erros) == 0 and len(log.sucessos) > 0:
                importacao.status = StatusImportacao.SUCESSO
            elif len(log.sucessos) > 0 and len(log.erros) > 0:
                importacao.status = StatusImportacao.SUCESSO_PARCIAL
            else:
                importacao.status = StatusImportacao.ERRO
                importacao.mensagem_erro = "Nenhum insumo foi importado com sucesso"

            self.db.commit()

            return True, f"Importação concluída: {len(log.sucessos)} sucessos, {len(log.erros)} erros"
        
        except Exception as e:
            logger.error(f"Erro ao processar importação: {e}")
            importacao.status = StatusImportacao.ERRO
            importacao.mensagem_erro = str(e)
            importacao.data_fim_processamento = datetime.now()
            self.db.commit()
            return False, f"Erro ao processar: {str(e)}"